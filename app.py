import streamlit as st
import openpyxl
import datetime
import io

# --- [설정: 만료 날짜] ---
EXPIRY_DATE = datetime.date(2026, 12, 31)

def check_license():
    """날짜가 지나면 서비스를 중단합니다."""
    current_date = datetime.date.today()
    if current_date > EXPIRY_DATE:
        st.error(f"⚠️ 서비스 이용 기간이 만료되었습니다. (만료일: {EXPIRY_DATE})")
        st.stop()

def process_excel(uploaded_file):
    """엑셀 파일의 수식을 값으로 변환합니다."""
    # 메모리 상에서 파일 읽기
    in_memory_file = io.BytesIO(uploaded_file.read())
    
    wb_data = openpyxl.load_workbook(in_memory_file, data_only=True)
    in_memory_file.seek(0)
    wb_origin = openpyxl.load_workbook(in_memory_file, data_only=False)

    for sheet_name in wb_origin.sheetnames:
        ws_data = wb_data[sheet_name]
        ws_origin = wb_origin[sheet_name]
        merged_ranges = ws_origin.merged_cells.ranges

        for row in ws_origin.iter_rows():
            for cell in row:
                coord = cell.coordinate
                is_merged = any(coord in mr for mr in merged_ranges)
                is_top_left = any(coord == mr.start_cell.coordinate for mr in merged_ranges if coord in mr)

                if not is_merged or is_top_left:
                    try:
                        cell.value = ws_data[coord].value
                    except:
                        pass
    
    # 결과를 메모리에 저장
    output = io.BytesIO()
    wb_origin.save(output)
    output.seek(0)
    return output

# --- 메인 웹 화면 구성 ---
st.set_page_config(page_title="엑셀 수식 제거기", page_icon="📊")
st.title("📊 엑셀 수식 → 값 변환 도구")
st.write(f"이 도구는 **{EXPIRY_DATE}**까지 사용 가능합니다.")

check_license()

uploaded_file = st.file_uploader("수식을 제거할 엑셀 파일을 업로드하세요.", type=["xlsx", "xlsm"])

if uploaded_file:
    with st.spinner("수식 제거 작업 중..."):
        try:
            processed_file = process_excel(uploaded_file)
            
            # 다운로드 버튼 생성
            st.success("✅ 변환 완료!")
            st.download_button(
                label="📁 변환된 파일 다운로드",
                data=processed_file,
                file_name=f"수식제거_{uploaded_file.name}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"오류 발생: {e}")
